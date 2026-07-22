from __future__ import annotations

import asyncio
import logging
import uuid
from typing import TYPE_CHECKING
from uuid import UUID

from app.domain.entities import Document
from app.domain.exceptions import (
    BulkLimitExceededError,
    ComputedVariableError,
    DocumentNotFoundError,
    InvalidSpreadsheetError,
    PdfConversionError,
    TemplateAccessDeniedError,
    TemplateVersionFileNotFoundError,
    TemplateVersionNotFoundError,
    VariablesMismatchError,
)
from app.domain.ports.document_repository import DocumentRepository
from app.domain.ports.storage_service import StorageService
from app.domain.ports.template_engine import TemplateEngine
from app.domain.ports.template_repository import TemplateRepository
from app.domain.services.computed_variables import computed_variable_names, resolve_computed
from app.domain.services.permissions import can_view_all_documents
from app.infrastructure.pdf.watermark import apply_watermark

if TYPE_CHECKING:
    from app.application.services.audit_service import AuditService
    from app.application.services.quota_service import QuotaService
    from app.application.services.usage_service import UsageService
    from app.domain.ports.pdf_converter import PdfConverter

logger = logging.getLogger(__name__)

# Concurrency insurance for the ephemeral preview endpoint ONLY — LibreOffice
# (behind Gotenberg) is CPU-heavy and this service runs on a 1vCPU host.
# Bounds preview() conversions to at most 2 concurrent regardless of how many
# preview requests land. The generate paths (generate_single, generate_bulk)
# are NOT gated by this semaphore — their concurrency is bounded separately
# by their own per-tier rate limits (see tier_limit_generate / tier_limit_bulk).
_PREVIEW_SEMAPHORE = asyncio.Semaphore(2)


class DocumentService:
    TEMPLATES_BUCKET = "templates"
    DOCUMENTS_BUCKET = "documents"

    def __init__(
        self,
        document_repository: DocumentRepository,
        template_repository: TemplateRepository,
        storage: StorageService,
        engine: TemplateEngine,
        pdf_converter: PdfConverter | None = None,
        bulk_generation_limit: int = 10,
        usage_service: UsageService | None = None,
        audit_service: AuditService | None = None,
        ip_address: str | None = None,
        quota_service: QuotaService | None = None,
        tier_id: UUID | None = None,
        user_bulk_override: int | None = None,
        preview_watermark_text: str | None = None,
    ):
        self._doc_repo = document_repository
        self._tpl_repo = template_repository
        self._storage = storage
        self._engine = engine
        self._pdf_converter = pdf_converter
        self._bulk_limit = bulk_generation_limit
        self._usage_service = usage_service
        self._audit_service = audit_service
        self._ip_address = ip_address
        self._quota_service = quota_service
        self._tier_id = tier_id
        self._user_bulk_override = user_bulk_override
        # Preview-only watermark text — falls back to Settings so callers
        # that don't wire it explicitly (e.g. tests) still get the
        # configured default. See get_document_service() for the explicit
        # DI wiring used in production.
        if preview_watermark_text is not None:
            self._preview_watermark_text = preview_watermark_text
        else:
            from app.config import get_settings

            self._preview_watermark_text = get_settings().preview_watermark_text

    async def generate_single(
        self,
        template_version_id: str,
        variables: dict[str, str],
        tenant_id: str,
        created_by: str,
        role: str = "user",
    ) -> dict:
        """
        Generate documents from a template version — the PRIMARY docx plus
        EVERY related file of the version, all rendered with the SAME
        resolved variable context:
        1. Fetch template version from DB
        2. Check user has access to the template
        3. Resolve computed variables ONCE
        4. Render + upload DOCX (and PDF) for the primary and each related file
        5. Create one document record per rendered file (primary first, then
           related files by position)
        6. Return {"documents": [Document, ...], "group_id": UUID | None}

        group_id is None when the version has no related files (single-file
        behavior unchanged); otherwise every document of this generation
        shares one uuid4.

        ATOMICITY: all-or-nothing across ALL files of the generation — any
        render/upload/PDF failure deletes every MinIO object uploaded so far
        and persists no DB rows.
        """
        # 1. Get template version
        version = await self._tpl_repo.get_version_by_id(uuid.UUID(template_version_id))
        if not version:
            raise TemplateVersionNotFoundError(
                f"Template version {template_version_id} not found"
            )

        # 2. Access check
        user_uuid = uuid.UUID(created_by)
        has_access = await self._tpl_repo.has_access(version.template_id, user_uuid, role)
        if not has_access:
            raise TemplateAccessDeniedError("No tenés acceso a esta plantilla")

        # 2a. Resolve template name for the enrichment fields (single lookup —
        # the FK is NOT NULL so the parent template always exists in prod).
        template = await self._tpl_repo.get_by_id(version.template_id)
        template_name = template.name if template is not None else ""

        # 2b. Resolve computed variables (server-authoritative) — AFTER
        # receiving the caller-supplied variables, BEFORE rendering, so both
        # the rendered documents and the persisted variables_snapshot below
        # reflect the resolved value. Resolved ONCE and reused for EVERY file.
        variables = resolve_computed(version.variables_meta, variables)

        related_files = sorted(
            list(getattr(version, "files", []) or []), key=lambda f: f.position
        )

        # 2c. Quota check (optional — skipped when quota_service is None).
        # Runs AFTER the related files are known because this generation
        # persists 1 + N documents (primary + every related file), and the
        # quota must be charged for ALL of them before anything is uploaded.
        if self._quota_service is not None and self._tier_id is not None:
            await self._quota_service.check_document_quota(
                tenant_id=uuid.UUID(tenant_id),
                tier_id=self._tier_id,
                additional=1 + len(related_files),
            )

        group_id: uuid.UUID | None = uuid.uuid4() if related_files else None

        # Base stem from the first variable value — the primary keeps the
        # historical naming exactly; related files inject their label before
        # the extension (spaces → underscores).
        first_var = next(iter(variables.values()), "") if variables else ""
        safe_name = "".join(
            c for c in str(first_var) if c.isalnum() or c in " _-"
        ).strip()[:50]

        # (label, source template path) — primary first, then files by position
        render_specs: list[tuple[str | None, str]] = [(None, version.minio_path)]
        render_specs += [(f.label, f.minio_path) for f in related_files]

        uploaded_paths: list[str] = []
        pending_documents: list[Document] = []

        try:
            for label, source_path in render_specs:
                # 3. Download the source template from MinIO
                template_bytes = await self._storage.download_file(
                    bucket=self.TEMPLATES_BUCKET,
                    path=source_path,
                )

                # 4. Render with docxtpl (same resolved context for every file)
                rendered_bytes = await self._engine.render(template_bytes, variables)

                doc_id = uuid.uuid4()
                if label is None:
                    docx_file_name = (
                        f"{safe_name}.docx" if safe_name else f"{doc_id}.docx"
                    )
                else:
                    stem = safe_name if safe_name else str(doc_id)
                    safe_label = (
                        "".join(c for c in label if c.isalnum() or c in " _-")
                        .strip()
                        .replace(" ", "_")
                    )
                    docx_file_name = f"{stem}_{safe_label or 'archivo'}.docx"

                docx_minio_path = f"{tenant_id}/{doc_id}/{docx_file_name}"
                await self._storage.upload_file(
                    bucket=self.DOCUMENTS_BUCKET,
                    path=docx_minio_path,
                    data=rendered_bytes,
                    content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                )
                uploaded_paths.append(docx_minio_path)

                # 4b. Convert DOCX → PDF (ADR-PDF-03: atomic dual-format)
                pdf_file_name: str | None = None
                pdf_minio_path: str | None = None

                if self._pdf_converter is not None:
                    pdf_bytes = await self._pdf_converter.convert(rendered_bytes)
                    pdf_stem = (
                        docx_file_name[:-5]
                        if docx_file_name.endswith(".docx")
                        else str(doc_id)
                    )
                    pdf_file_name = f"{pdf_stem}.pdf"
                    pdf_minio_path = f"{tenant_id}/{doc_id}/{pdf_file_name}"
                    await self._storage.upload_file(
                        bucket=self.DOCUMENTS_BUCKET,
                        path=pdf_minio_path,
                        data=pdf_bytes,
                        content_type="application/pdf",
                    )
                    uploaded_paths.append(pdf_minio_path)

                # Build domain entity (not persisted until ALL files succeed)
                pending_documents.append(
                    Document(
                        id=doc_id,
                        tenant_id=uuid.UUID(tenant_id),
                        template_version_id=uuid.UUID(template_version_id),
                        docx_minio_path=docx_minio_path,
                        docx_file_name=docx_file_name,
                        pdf_file_name=pdf_file_name,
                        pdf_minio_path=pdf_minio_path,
                        generation_type="single",
                        group_id=group_id,
                        variables_snapshot=variables,
                        created_by=uuid.UUID(created_by),
                        status="completed",
                        template_id=version.template_id,
                        template_name=template_name,
                        template_version=version.version,
                    )
                )
        except Exception:
            # Atomic rollback: delete EVERY object uploaded so far (best-effort)
            for path in uploaded_paths:
                try:
                    await self._storage.delete_file(self.DOCUMENTS_BUCKET, path)
                except Exception:
                    logger.warning(
                        "Failed to delete upload %s during single-generation rollback",
                        path,
                    )
            raise  # re-raise so the caller sees the original error

        # 5. Create DB records (all files succeeded — repo enriches each row)
        documents = [await self._doc_repo.create(d) for d in pending_documents]

        # 6. Record usage + audit (both optional for backward compat)
        user_uuid = uuid.UUID(created_by)
        tenant_uuid = uuid.UUID(tenant_id)
        template_uuid = version.template_id

        if self._usage_service is not None:
            await self._usage_service.record(
                user_id=user_uuid,
                tenant_id=tenant_uuid,
                template_id=template_uuid,
                generation_type="single",
                document_count=len(documents),
            )

        if self._audit_service is not None:
            from app.domain.entities import AuditAction
            audit_details: dict = {}
            if self._pdf_converter is not None:
                audit_details["formats_generated"] = ["docx", "pdf"]
            if group_id is not None:
                audit_details["document_count"] = len(documents)
                audit_details["group_id"] = str(group_id)
            self._audit_service.log(
                actor_id=user_uuid,
                tenant_id=tenant_uuid,
                action=AuditAction.DOCUMENT_GENERATE,
                resource_type="document",
                resource_id=documents[0].id,
                details=audit_details if audit_details else None,
                ip_address=self._ip_address,
            )

        return {
            "documents": documents,
            "group_id": group_id,
        }

    async def preview(
        self,
        *,
        template_version_id: str,
        variables: dict[str, str],
        user_id: str,
        role: str = "user",
        file_id: str | None = None,
    ) -> bytes:
        """
        Render a template with the CURRENT (possibly partial) variable
        values and return the converted PDF bytes directly. Nothing is
        persisted: no MinIO uploads, no Document row, no usage/audit
        tracking, no quota check.

        When `file_id` is given, the RELATED file of the version is
        previewed instead of the primary docx (same variable context;
        watermark unchanged).

        Missing variables render as blanks — this is the default Jinja2
        Undefined behavior in docxtpl, so no extra handling is needed here.

        Uses the same collaborators and access check as generate_single()
        so version-not-found / access-denied map to identical exceptions.

        Raises:
            TemplateVersionNotFoundError: same as generate_single.
            TemplateVersionFileNotFoundError: file_id doesn't exist OR belongs
                to a different version.
            TemplateAccessDeniedError: same as generate_single.
            PdfConversionError: propagated from the converter, uncaught.
        """
        # 1. Get template version — same check as generate_single
        version = await self._tpl_repo.get_version_by_id(uuid.UUID(template_version_id))
        if not version:
            raise TemplateVersionNotFoundError(
                f"Template version {template_version_id} not found"
            )

        # 2. Access check — same check as generate_single
        user_uuid = uuid.UUID(user_id)
        has_access = await self._tpl_repo.has_access(version.template_id, user_uuid, role)
        if not has_access:
            raise TemplateAccessDeniedError("No tenés acceso a esta plantilla")

        # 2b. Resolve computed variables — same server-authoritative rule as
        # generate_single. Missing/unparseable sources resolve to "" so a
        # partially-filled preview never crashes.
        variables = resolve_computed(version.variables_meta, variables)

        # 2c. Optional related-file selection — preview that file's docx
        # instead of the primary (must belong to this version).
        source_path = version.minio_path
        if file_id is not None:
            file = await self._tpl_repo.get_version_file(
                uuid.UUID(template_version_id), uuid.UUID(file_id)
            )
            if file is None:
                raise TemplateVersionFileNotFoundError(
                    f"Template version file {file_id} not found"
                )
            source_path = file.minio_path

        # 3. Download template from MinIO (read-only — no upload_file calls below)
        template_bytes = await self._storage.download_file(
            bucket=self.TEMPLATES_BUCKET,
            path=source_path,
        )

        # 4. Render with the current (possibly partial) variables, verbatim
        rendered = await self._engine.render(template_bytes, variables)

        # 5. Convert to PDF, guarded by a module-level semaphore (concurrency
        # insurance for LibreOffice on a 1vCPU host). PdfConversionError
        # propagates uncaught — there is nothing to roll back (ephemeral).
        if self._pdf_converter is None:
            raise PdfConversionError("PdfConverter not configured — cannot preview")

        async with _PREVIEW_SEMAPHORE:
            pdf_bytes = await self._pdf_converter.convert(rendered)

        # 6. Stamp a diagonal semi-transparent watermark on every page so
        # this preview can never be mistaken for (or used as) a final
        # document. Watermarking is local CPU work — it runs AFTER the
        # semaphore is released so it never holds a Gotenberg conversion
        # slot. Offloaded to a thread since apply_watermark is sync/CPU-bound.
        pdf_bytes = await asyncio.to_thread(
            apply_watermark, pdf_bytes, self._preview_watermark_text
        )

        return pdf_bytes

    def _can_access_document(
        self, document: Document, *, requester_id: uuid.UUID, role: str
    ) -> bool:
        """Ownership rule shared by get / download / delete / batch (finding #1).

        Admins (can_view_all_documents) may access ANY document; every other
        role may access ONLY documents they created. This is the exact same
        decision the LIST endpoint applies via `created_by` scoping — kept in
        one place so detail/download/delete/batch can never drift from it.
        """
        return can_view_all_documents(role) or document.created_by == requester_id

    async def get_document(
        self, document_id: uuid.UUID, *, requester_id: uuid.UUID, role: str
    ) -> dict:
        """Get document by ID with fresh presigned URL.

        Enforces ownership (finding #1): a non-admin requester that did not
        create the document gets a non-leaking DocumentNotFoundError (mapped
        to 404), mirroring the folder/preset "not yours -> NotFound"
        convention so the response never confirms the document's existence.
        """
        document = await self._doc_repo.get_by_id(document_id)
        if not document or not self._can_access_document(
            document, requester_id=requester_id, role=role
        ):
            raise DocumentNotFoundError(f"Document {document_id} not found")

        download_url = await self._storage.get_presigned_url(
            bucket=self.DOCUMENTS_BUCKET,
            path=document.docx_minio_path,
        )

        return {
            "document": document,
            "download_url": download_url,
        }

    async def download_document(self, minio_path: str) -> bytes:
        """Download a generated document from MinIO.

        Access is gated upstream: callers resolve the document via
        get_document() (which enforces ownership) before requesting bytes by
        path.
        """
        return await self._storage.download_file(
            bucket=self.DOCUMENTS_BUCKET,
            path=minio_path,
        )

    async def delete_document(
        self, document_id: uuid.UUID, *, requester_id: uuid.UUID, role: str
    ) -> None:
        """Delete a document record and its files from MinIO.

        Enforces ownership (finding #1): a non-admin requester that did not
        create the document gets a non-leaking DocumentNotFoundError and the
        document is left untouched. Both the DOCX and the PDF objects are
        removed (finding #4) so dual-format documents never orphan a PDF.
        """
        document = await self._doc_repo.get_by_id(document_id)
        if not document or not self._can_access_document(
            document, requester_id=requester_id, role=role
        ):
            raise DocumentNotFoundError(f"Document {document_id} not found")
        # Delete both MinIO objects (best-effort — a file may already be gone).
        # The PDF was previously left behind, orphaning it on every delete.
        for path in (document.docx_minio_path, document.pdf_minio_path):
            if not path:
                continue
            try:
                await self._storage.delete_file(self.DOCUMENTS_BUCKET, path)
            except Exception:
                pass  # File may already be gone
        # Delete from DB
        await self._doc_repo.delete(document_id)

        # Audit the deletion (optional for backward compat)
        if self._audit_service is not None:
            from app.domain.entities import AuditAction
            self._audit_service.log(
                actor_id=document.created_by,
                tenant_id=document.tenant_id,
                action=AuditAction.DOCUMENT_DELETE,
                resource_type="document",
                resource_id=document_id,
                ip_address=self._ip_address,
            )

    async def get_template_owner_id(self, template_id: str) -> uuid.UUID | None:
        """Return the owner (created_by) UUID for the given template_id string.

        Returns None if the template does not exist.
        Used by the documents list endpoint to bypass the created_by filter
        for the template's owner (REQ-OWN-DOCS).
        """
        return await self._tpl_repo.get_owner_id(uuid.UUID(template_id))

    async def list_documents(
        self,
        page: int = 1,
        size: int = 20,
        template_id: str | None = None,
        created_by: str | None = None,
    ) -> tuple[list, int]:
        """List documents with pagination."""
        tpl_uuid = uuid.UUID(template_id) if template_id else None
        created_by_uuid = uuid.UUID(created_by) if created_by else None
        return await self._doc_repo.list_paginated(
            page=page, size=size, template_id=tpl_uuid, created_by=created_by_uuid
        )

    async def list_documents_by_batch(
        self, batch_id: UUID, tenant_id: UUID, *, requester_id: UUID, role: str
    ) -> list:
        """Return the documents for a given batch, scoped to tenant AND owner.

        Public delegator for DocumentRepository.list_by_batch_id.
        Replaces the private _doc_repo access pattern used in the bulk
        download endpoint (W-PRES-02 fix). O(batch_size) instead of O(N total).

        Ownership (finding #1): a non-admin requester only sees documents they
        created. A whole batch is created by a single user, so a non-creator
        non-admin gets an empty list — the endpoint then returns a non-leaking
        404 exactly as it does for a non-existent batch. Admins see everything.
        """
        docs = await self._doc_repo.list_by_batch_id(
            batch_id=batch_id, tenant_id=tenant_id
        )
        if can_view_all_documents(role):
            return docs
        return [d for d in docs if d.created_by == requester_id]

    # ── Bulk generation methods ──────────────────────────────────────────

    async def generate_excel_template(
        self,
        template_version_id: str,
        user_id: str | None = None,
        role: str = "user",
    ) -> tuple[bytes, str]:
        """
        Generate a blank Excel template for bulk data entry.
        Column headers = template variable names.
        Returns (excel_bytes, filename).
        """
        import openpyxl
        from io import BytesIO
        from openpyxl.styles import Alignment, Font, PatternFill

        version = await self._tpl_repo.get_version_by_id(
            uuid.UUID(template_version_id)
        )
        if not version:
            raise TemplateVersionNotFoundError(
                f"Template version {template_version_id} not found"
            )

        if user_id is not None:
            user_uuid = uuid.UUID(user_id)
            has_access = await self._tpl_repo.has_access(
                version.template_id, user_uuid, role
            )
            if not has_access:
                raise TemplateAccessDeniedError("No tenés acceso a esta plantilla")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # Computed variables are server-resolved, never supplied by the
        # user — exclude them from the bulk data-entry columns.
        computed_names = computed_variable_names(version.variables_meta)
        variables = [v for v in version.variables if v not in computed_names]

        # Header row with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", fill_type="solid")

        for col, var_name in enumerate(variables, 1):
            cell = ws.cell(row=1, column=col, value=var_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(
                len(var_name) + 4, 15
            )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Build filename from template name
        template = await self._tpl_repo.get_by_id(version.template_id)
        safe_name = "".join(
            c for c in template.name if c.isalnum() or c in " _-"
        ).strip()
        filename = f"{safe_name}_bulk_template.xlsx"

        return output.read(), filename

    async def parse_excel_data(
        self,
        template_version_id: str,
        excel_bytes: bytes,
        user_id: str | None = None,
        role: str = "user",
        tenant_id: str | None = None,
        user_bulk_override: int | None = None,
    ) -> list[dict[str, str]]:
        """
        Parse a filled Excel file and validate against template variables.
        Returns list of dicts (one per row).
        Raises BulkLimitExceededError if > bulk_limit rows (when quota_service is None).
        Raises QuotaExceededError if > tier/user bulk limit (when quota_service is set).
        Raises VariablesMismatchError if headers don't match template variables.
        Raises TemplateAccessDeniedError if user lacks access.
        """
        import openpyxl
        import zipfile
        from io import BytesIO

        from openpyxl.utils.exceptions import InvalidFileException

        version = await self._tpl_repo.get_version_by_id(
            uuid.UUID(template_version_id)
        )
        if not version:
            raise TemplateVersionNotFoundError(
                f"Template version {template_version_id} not found"
            )

        if user_id is not None:
            user_uuid = uuid.UUID(user_id)
            has_access = await self._tpl_repo.has_access(
                version.template_id, user_uuid, role
            )
            if not has_access:
                raise TemplateAccessDeniedError("No tenés acceso a esta plantilla")

        # Parsing untrusted upload bytes. A .docx renamed to .xlsx passes the
        # endpoint's extension check but is not a readable workbook, and raw
        # non-zip bytes fail even earlier. openpyxl surfaces these as
        # zipfile.BadZipFile / KeyError / InvalidFileException — none of which
        # are domain errors, so left unwrapped they escaped as an uncaught 500.
        # Map them to a domain error the API layer turns into a clean 400.
        try:
            wb = openpyxl.load_workbook(BytesIO(excel_bytes))
            ws = wb.active
            # Read headers from first row
            headers = [cell.value for cell in ws[1] if cell.value is not None]
        except (zipfile.BadZipFile, KeyError, InvalidFileException) as exc:
            raise InvalidSpreadsheetError(
                "El archivo no es un Excel (.xlsx) válido o está dañado. "
                "Descargá la plantilla de ejemplo y volvé a intentarlo."
            ) from exc

        # Reject duplicate header columns explicitly. The mismatch check below
        # is set-based, so two columns with the same name would silently
        # collapse (last-write-wins) instead of failing — a data-loss trap.
        seen: set = set()
        duplicates: list = []
        for header in headers:
            if header in seen and header not in duplicates:
                duplicates.append(header)
            seen.add(header)
        if duplicates:
            joined = ", ".join(str(d) for d in duplicates)
            raise VariablesMismatchError(
                f"El archivo tiene columnas duplicadas: {joined}. "
                "Cada variable debe aparecer una sola vez."
            )

        # Validate headers match template variables. Computed variables are
        # server-resolved, never supplied by the user — excluded from the
        # expected set so bulk uploads don't need (and can't provide) a
        # column for them.
        computed_names = computed_variable_names(version.variables_meta)
        expected = set(version.variables) - computed_names
        actual = set(headers)
        if expected != actual:
            missing = expected - actual
            extra = actual - expected
            raise VariablesMismatchError(
                f"Header mismatch. Missing: {missing or 'none'}. "
                f"Extra: {extra or 'none'}"
            )

        # Parse data rows
        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            row_dict = {}
            for header, value in zip(headers, row):
                row_dict[header] = str(value) if value is not None else ""
            rows.append(row_dict)

        if len(rows) == 0:
            raise VariablesMismatchError("Excel file has no data rows")

        # Absolute hard cap — ALWAYS enforced, independent of the quota flag or
        # whether tenant_id was threaded through (finding #2). This is the DoS
        # guardrail: production wires quota_service + tier_id but never passes
        # tenant_id, so the quota branch below used to skip the check entirely
        # while the legacy fallback never ran — leaving bulk requests uncapped.
        # `self._bulk_limit` already reflects the effective per-user/global
        # limit (see get_document_service). Rejected BEFORE any row is rendered.
        if len(rows) > self._bulk_limit:
            raise BulkLimitExceededError(limit=self._bulk_limit)

        # Tier/user quota bulk limit (defense-in-depth). Only ever further
        # restricts the request; it can never loosen the absolute cap above.
        if self._quota_service is not None and self._tier_id is not None:
            effective_tenant_id = uuid.UUID(tenant_id) if tenant_id else None
            if effective_tenant_id is not None:
                await self._quota_service.check_bulk_limit(
                    tenant_id=effective_tenant_id,
                    tier_id=self._tier_id,
                    requested_count=len(rows),
                    user_bulk_override=user_bulk_override,
                )

        return rows

    async def generate_bulk(
        self,
        template_version_id: str,
        rows: list[dict[str, str]],
        tenant_id: str,
        created_by: str,
        role: str = "user",
    ) -> dict:
        """
        Generate multiple documents from template + data rows.
        Returns dict with batch_id, zip_path, document_count, and errors.
        """
        import zipfile
        from io import BytesIO

        version = await self._tpl_repo.get_version_by_id(
            uuid.UUID(template_version_id)
        )
        if not version:
            raise TemplateVersionNotFoundError(
                f"Template version {template_version_id} not found"
            )

        # Access check
        user_uuid = uuid.UUID(created_by)
        has_access = await self._tpl_repo.has_access(version.template_id, user_uuid, role)
        if not has_access:
            raise TemplateAccessDeniedError("No tenés acceso a esta plantilla")

        # Related files of the version: downloaded ONCE (below), rendered for
        # EVERY row with the same (resolved) row context as the primary.
        related_files = sorted(
            list(getattr(version, "files", []) or []), key=lambda f: f.position
        )

        # Quota checks (optional — skipped when quota_service is None). Run
        # AFTER the related files are known: every row persists 1 + N
        # documents (primary + related files), so the document quota must be
        # charged rows × (1 + N) before anything is uploaded.
        if self._quota_service is not None and self._tier_id is not None:
            tenant_uuid = uuid.UUID(tenant_id)
            await self._quota_service.check_document_quota(
                tenant_id=tenant_uuid,
                tier_id=self._tier_id,
                additional=len(rows) * (1 + len(related_files)),
            )
            await self._quota_service.check_bulk_limit(
                tenant_id=tenant_uuid,
                tier_id=self._tier_id,
                requested_count=len(rows),
                user_bulk_override=self._user_bulk_override,
            )

        # Resolve template name ONCE for the enrichment fields on every row
        template = await self._tpl_repo.get_by_id(version.template_id)
        template_name = template.name if template is not None else ""

        # Download template ONCE — engine creates a fresh DocxTemplate per render
        template_bytes = await self._storage.download_file(
            bucket=self.TEMPLATES_BUCKET,
            path=version.minio_path,
        )

        related_specs: list[tuple[str, bytes]] = []
        for related in related_files:
            related_bytes = await self._storage.download_file(
                bucket=self.TEMPLATES_BUCKET,
                path=related.minio_path,
            )
            related_specs.append((related.label, related_bytes))

        batch_id = uuid.uuid4()
        zip_buffer = BytesIO()
        documents = []
        # Track all MinIO keys uploaded so far for rollback on failure (ADR-PDF-03 bulk)
        uploaded_minio_paths: list[str] = []

        # Finding #3: wrap the ENTIRE upload + persist region so that ANY
        # exception after the first upload — a render failure, an upload
        # failure, a ComputedVariableError, a PdfConversionError, or a failure
        # while persisting the batch to the DB — best-effort deletes every
        # object uploaded so far (docx/pdf AND the batch ZIP) and re-raises.
        # Previously only ComputedVariableError and PdfConversionError rolled
        # back; every other failure orphaned already-uploaded objects (no DB
        # rows -> a retry would duplicate them).
        try:
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for i, row_data in enumerate(rows):
                    # Resolve computed variables (server-authoritative) — same
                    # rule as generate_single/preview, applied per row so both
                    # the rendered document and the persisted variables_snapshot
                    # reflect resolved values. A ComputedVariableError mid-batch
                    # is caught by the outer handler and rolls the batch back.
                    row_data = resolve_computed(version.variables_meta, row_data)

                    # Generate the row's base filename from its first variable value
                    first_var = next(iter(row_data.values()), f"doc_{i + 1}")
                    safe_name = "".join(
                        c for c in str(first_var) if c.isalnum() or c in " _-"
                    ).strip()[:50]
                    base_docx_file_name = (
                        f"{i + 1:03d}_{safe_name}.docx"
                        if safe_name
                        else f"{i + 1:03d}_document.docx"
                    )

                    # group_id shared PER ROW when the version has related files
                    row_group_id: uuid.UUID | None = (
                        uuid.uuid4() if related_specs else None
                    )

                    # Primary first, then related files by position — the label
                    # is injected before the extension for related files.
                    file_specs: list[tuple[str | None, bytes]] = [(None, template_bytes)]
                    file_specs += related_specs

                    for label, source_bytes in file_specs:
                        # Render document (fresh DocxTemplate per render — engine)
                        rendered_bytes = await self._engine.render(source_bytes, row_data)

                        if label is None:
                            docx_file_name = base_docx_file_name
                        else:
                            safe_label = (
                                "".join(c for c in label if c.isalnum() or c in " _-")
                                .strip()
                                .replace(" ", "_")
                            )
                            stem = (
                                base_docx_file_name[:-5]
                                if base_docx_file_name.endswith(".docx")
                                else base_docx_file_name
                            )
                            docx_file_name = f"{stem}_{safe_label or 'archivo'}.docx"

                        zf.writestr(docx_file_name, rendered_bytes)

                        doc_id = uuid.uuid4()
                        docx_minio_path = f"{tenant_id}/{batch_id}/{docx_file_name}"

                        # Upload DOCX to MinIO
                        await self._storage.upload_file(
                            bucket=self.DOCUMENTS_BUCKET,
                            path=docx_minio_path,
                            data=rendered_bytes,
                            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        )
                        uploaded_minio_paths.append(docx_minio_path)

                        # Convert DOCX → PDF (ADR-PDF-03 bulk: atomic, sequential)
                        pdf_file_name: str | None = None
                        pdf_minio_path: str | None = None

                        if self._pdf_converter is not None:
                            pdf_bytes = await self._pdf_converter.convert(rendered_bytes)

                            # Upload PDF to MinIO
                            pdf_stem = (
                                docx_file_name[:-5]
                                if docx_file_name.endswith(".docx")
                                else f"{doc_id}"
                            )
                            pdf_file_name = f"{pdf_stem}.pdf"
                            pdf_minio_path = f"{tenant_id}/{batch_id}/{pdf_file_name}"
                            await self._storage.upload_file(
                                bucket=self.DOCUMENTS_BUCKET,
                                path=pdf_minio_path,
                                data=pdf_bytes,
                                content_type="application/pdf",
                            )
                            uploaded_minio_paths.append(pdf_minio_path)

                        # Build domain entity (not persisted until all rows succeed)
                        doc = Document(
                            id=doc_id,
                            tenant_id=uuid.UUID(tenant_id),
                            template_version_id=uuid.UUID(template_version_id),
                            docx_minio_path=docx_minio_path,
                            docx_file_name=docx_file_name,
                            pdf_file_name=pdf_file_name,
                            pdf_minio_path=pdf_minio_path,
                            generation_type="bulk",
                            batch_id=batch_id,
                            group_id=row_group_id,
                            variables_snapshot=row_data,
                            created_by=uuid.UUID(created_by),
                            status="completed",
                            template_id=version.template_id,
                            template_name=template_name,
                            template_version=version.version,
                        )
                        documents.append(doc)

            zip_buffer.seek(0)
            zip_bytes = zip_buffer.read()

            # Upload ZIP to MinIO
            zip_path = f"{tenant_id}/{batch_id}/bulk.zip"
            await self._storage.upload_file(
                bucket=self.DOCUMENTS_BUCKET,
                path=zip_path,
                data=zip_bytes,
                content_type="application/zip",
            )
            # Track the ZIP too so a DB-persist failure below rolls it back.
            uploaded_minio_paths.append(zip_path)

            # Save document records to DB (all rows succeeded — persist atomically)
            if documents:
                await self._doc_repo.create_batch(documents)
        except Exception:
            # Any failure after the first upload: delete EVERY uploaded object
            # (docx/pdf/zip) best-effort, then re-raise so no partial batch and
            # no orphaned MinIO objects survive.
            for path in uploaded_minio_paths:
                try:
                    await self._storage.delete_file(self.DOCUMENTS_BUCKET, path)
                except Exception:
                    logger.warning(
                        "Failed to delete bulk upload %s during rollback",
                        path,
                    )
            raise  # propagate — no DB rows persisted

        # Record usage + audit (both optional for backward compat)
        success_count = len(documents)
        user_uuid = uuid.UUID(created_by)
        tenant_uuid = uuid.UUID(tenant_id)
        template_uuid = version.template_id

        if self._usage_service is not None and success_count > 0:
            await self._usage_service.record(
                user_id=user_uuid,
                tenant_id=tenant_uuid,
                template_id=template_uuid,
                generation_type="bulk",
                document_count=success_count,
            )

        if self._audit_service is not None:
            from app.domain.entities import AuditAction
            audit_details: dict = {
                "document_count": success_count,
            }
            if self._pdf_converter is not None:
                audit_details["formats_generated"] = ["docx", "pdf"]
            self._audit_service.log(
                actor_id=user_uuid,
                tenant_id=tenant_uuid,
                action=AuditAction.DOCUMENT_GENERATE_BULK,
                resource_type="document_batch",
                resource_id=batch_id,
                details=audit_details,
                ip_address=self._ip_address,
            )

        return {
            "batch_id": batch_id,
            "zip_path": zip_path,
            "document_count": success_count,
            "errors": [],
        }

    # ── Lazy PDF backfill ───────────────────────────────────────────────────

    async def ensure_pdf(self, document_id: UUID) -> Document:
        """Lazily generate and persist a PDF for a legacy DOCX-only document.

        ADR-PDF-04 contract:
        - Fast path (idempotent): if pdf_file_name is already set, return
          the document immediately without calling the converter.
        - Slow path: download DOCX bytes, convert to PDF, upload PDF, update
          document row via update_pdf_fields(), return updated document.
        - On PdfConversionError: do NOT delete DOCX, do NOT update DB, let
          exception propagate so the presentation layer maps it to HTTP 503.

        Concurrency note: two concurrent requests for the same legacy doc may
        both run the slow path (last-write-wins). The PDF bytes are
        deterministic (same DOCX input → same PDF output for a given Gotenberg
        version), so duplicate uploads are tolerable and the row ends up in a
        consistent state either way.

        Returns:
            The updated Document with pdf_file_name and pdf_minio_path set.

        Raises:
            DocumentNotFoundError: if document_id does not exist.
            PdfConversionError: if converter fails (pdf_file_name remains NULL).
        """
        document = await self._doc_repo.get_by_id(document_id)
        if document is None:
            raise DocumentNotFoundError(f"Document {document_id} not found")

        # Fast path — already has PDF (idempotent)
        if document.pdf_file_name is not None:
            return document

        # Slow path — backfill
        if self._pdf_converter is None:
            raise PdfConversionError(
                "PdfConverter not configured — cannot backfill PDF"
            )

        # Download DOCX from MinIO (raises FileNotFoundError on missing file)
        docx_bytes = await self._storage.download_file(
            bucket=self.DOCUMENTS_BUCKET,
            path=document.docx_minio_path,
        )

        # Convert — propagate PdfConversionError without touching DB or MinIO
        pdf_bytes = await self._pdf_converter.convert(docx_bytes)

        # Derive PDF path from DOCX path (deterministic key for idempotency)
        docx_minio_path = document.docx_minio_path
        if docx_minio_path.endswith(".docx"):
            pdf_minio_path = docx_minio_path[:-5] + ".pdf"
        else:
            pdf_minio_path = docx_minio_path + ".pdf"

        pdf_file_name = (
            document.docx_file_name[:-5] + ".pdf"
            if document.docx_file_name.endswith(".docx")
            else document.docx_file_name + ".pdf"
        )

        # Upload PDF to MinIO
        await self._storage.upload_file(
            bucket=self.DOCUMENTS_BUCKET,
            path=pdf_minio_path,
            data=pdf_bytes,
            content_type="application/pdf",
        )

        # Persist update (S-02 resolution: update_pdf_fields returns a domain
        # Document so callers do not need to handle ORM objects directly)
        updated = await self._doc_repo.update_pdf_fields(
            doc_id=document_id,
            pdf_file_name=pdf_file_name,
            pdf_minio_path=pdf_minio_path,
        )

        # Both the real repo and the fake return an updated domain Document
        # (the real repo re-fetches with eager-loaded relationships so the
        # template enrichment fields are populated too).
        return updated

    # ------------------------------------------------------------------
    # W-PRES-03: public audit helpers — presentation layer calls these
    # instead of reaching into self._audit_service directly.
    # ------------------------------------------------------------------

    async def log_download_event(
        self,
        *,
        actor_id: UUID,
        tenant_id: UUID,
        document_id: UUID,
        format: str,
        via: str,
    ) -> None:
        """Record a single-document download in the audit log.

        Delegates to self._audit_service.log() if audit_service is set.
        No-op when audit_service is None (backward compat).
        """
        if self._audit_service is None:
            return
        from app.domain.entities import AuditAction
        self._audit_service.log(
            actor_id=actor_id,
            tenant_id=tenant_id,
            action=AuditAction.DOCUMENT_DOWNLOAD,
            resource_type="document",
            resource_id=document_id,
            details={
                "format": format,
                "document_id": str(document_id),
                "via": via,
            },
            ip_address=None,
        )

    async def log_bulk_download_event(
        self,
        *,
        actor_id: UUID,
        tenant_id: UUID,
        batch_id: UUID,
        format: str,
        via: str,
        include_both: bool,
    ) -> None:
        """Record a bulk-document download in the audit log.

        Delegates to self._audit_service.log() if audit_service is set.
        No-op when audit_service is None (backward compat).
        """
        if self._audit_service is None:
            return
        from app.domain.entities import AuditAction
        self._audit_service.log(
            actor_id=actor_id,
            tenant_id=tenant_id,
            action=AuditAction.DOCUMENT_DOWNLOAD,
            resource_type="document_batch",
            resource_id=batch_id,
            details={
                "format": format,
                "document_id": str(batch_id),
                "via": via,
                "include_both": include_both,
            },
            ip_address=None,
        )
